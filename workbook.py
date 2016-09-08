from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl import Workbook
from operator import itemgetter
from collections import OrderedDict


class PlayerWorkbook():

    yellow_fill = PatternFill(patternType='solid', fgColor=Color("E6F2A0"))
    red_fill = PatternFill(patternType='solid', fgColor=Color("D49093"))
    blue_fill = PatternFill(patternType='solid', fgColor=Color("84D6F0"))
    orange_fill = PatternFill(patternType='solid', fgColor=Color("FCB082"))
    pink_fill = PatternFill(patternType='solid', fgColor=Color("FADCF4"))
    center_alignment = Alignment(horizontal='center')

    def __init__(self, league_gap_threshold):
        self.league_gap_threshold = league_gap_threshold

    def build_workbook(self, leagues):
        self.players_wb = Workbook()
        leagues = sorted(leagues, key=itemgetter('league'))
        self._write_players_sheet(leagues)
        for league in leagues:
            self._write_league_sheet(league)
        return self.players_wb

    def _write_players_sheet(self, leagues):
        ws = self.players_wb.worksheets[0]
        self._write_playersheet_headers(ws, leagues)
        self._write_playersheet_timelines(ws, leagues)
        return ws

    @staticmethod
    def _players_from_leagues(leagues):
        players = set()
        for league in leagues:
            for player in league['players'].keys():
                players.add(player)
        players = sorted(players)
        return players

    def _write_playersheet_headers(self, worksheet, leagues):
        worksheet.title = "Players"
        worksheet['A2'].value = "Name"
        worksheet['B2'].value = "Longest Gap"
        worksheet['C2'].value = "Weeks Played"
        worksheet['A2'].fill = self.blue_fill
        worksheet['C2'].fill = self.blue_fill
        worksheet['B2'].fill = self.blue_fill

        # Create header for each league
        for col, league in enumerate(leagues, start=1):
            league_num = league['league']
            league_header_cell = worksheet.cell(row=2, column=col+3)
            league_header_cell.value = league_num
            league_header_cell.fill = self.blue_fill
            league_header_cell.alignment = self.center_alignment
            worksheet.column_dimensions[league_header_cell.column].width = 3
            last_column = column_index_from_string(league_header_cell.column)

        # Create title to span leagues
        worksheet.merge_cells(start_row=1, start_column=3,
                              end_row=1, end_column=last_column)
        league_title_cell = worksheet.cell(row=1, column=3)
        league_title_cell.value = "Leagues"
        league_title_cell.alignment = self.center_alignment
        league_title_cell.fill = self.orange_fill

    def _write_playersheet_timelines(self, worksheet, leagues):
        """
        For each player, display how many times that name appeared for each league.
        Also display the biggest gap a player has between leagues
        """
        players = self._players_from_leagues(leagues)
        for league, player in enumerate(players, start=2):
            row = 1 + league
            player_cell = worksheet.cell(row=row, column=1)
            biggest_league_gap_cell = worksheet.cell(row=row, column=2)
            week_count_cell = worksheet.cell(row=row, column=3)
            player_cell.value = player
            player_clash = False
            started_bowling = False
            week_count = 0
            league_gap = 0
            league_gaps = []
            for col, league in enumerate(leagues, start=1):
                league_cell = worksheet.cell(row=row, column=col+3)
                if player not in league['players']:
                    if started_bowling:  # Proving the player has bowled before
                        league_gap += 1
                    continue
                if started_bowling and (league_gap >= self.league_gap_threshold):
                    for previous_col in range(league_gap):
                        pc = worksheet.cell(row=row, column=col+3-previous_col-1)
                        pc.fill = self.pink_fill
                    biggest_league_gap_cell.fill = self.pink_fill
                league_gaps.append(league_gap)
                league_gap = 0
                started_bowling = True
                weeks = league['players'][player]
                played_weeks = sum(weeks.values())
                league_cell.value = played_weeks
                league_cell.alignment = self.center_alignment
                max_key = max(weeks, key=weeks.get)
                # If a name appears more than once in any one week, thats bad.
                # ...so mark that cell as red
                if weeks[max_key] == 1:
                    league_cell.fill = self.yellow_fill
                elif weeks[max_key] > 1:
                    league_cell.fill = self.red_fill
                    player_clash = True
                week_count += played_weeks
            week_count_cell.value = week_count
            biggest_league_gap_cell.value = max(league_gaps)
            if player_clash:
                week_count_cell.fill = self.red_fill

    def _write_leaguesheet_headers(self, worksheet, no_of_weeks):
        worksheet['A2'].value = "Name"
        worksheet['B2'].value = "Weeks Played"
        worksheet['A2'].fill = self.blue_fill
        worksheet['B2'].fill = self.blue_fill
        for week in range(1, no_of_weeks+1):
            week_header_cell = worksheet.cell(row=2, column=week+2)
            week_header_cell.value = week
            week_header_cell.fill = self.blue_fill
            week_header_cell.alignment = self.center_alignment
            worksheet.column_dimensions[week_header_cell.column].width = 3
            last_column = column_index_from_string(week_header_cell.column)
        worksheet.merge_cells(
            start_row=1, start_column=3, end_row=1, end_column=last_column
        )
        week_title_cell = worksheet.cell(row=1, column=3)
        week_title_cell.value = "Weeks"
        week_title_cell.fill = self.orange_fill
        week_title_cell.alignment = self.center_alignment

    def _write_league_sheet(self, league):
        worksheet = self.players_wb.create_sheet()
        worksheet.title = "leauge {}".format(league['league'])
        no_of_weeks = league['no_of_weeks']
        self._write_leaguesheet_headers(worksheet, no_of_weeks)
        self._write_playerweekly_timelines(worksheet, league)

    def _write_playerweekly_timelines(self, worksheet, league):
        """
        For each player, display how many times that name appeared for each week.
        """
        players = OrderedDict(sorted(league['players'].items()))
        for num, (player, weeks) in enumerate(players.items(), start=2):
            row = num + 1
            worksheet.cell(row=row, column=1).value = player
            worksheet.cell(row=row, column=2).value = sum(weeks.values())
            for week, count in weeks.items():
                weekcell = worksheet.cell(row=row, column=week+2)
                weekcell.value = count
                weekcell.alignment = self.center_alignment
                if count == 1:
                    weekcell.fill = self.yellow_fill
                elif count > 1:
                    weekcell.fill = self.red_fill
