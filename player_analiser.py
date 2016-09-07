from openpyxl import Workbook
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from operator import itemgetter
from openpyxl.utils import column_index_from_string
import sys
import glob
import re
import os
from loadplayers import load_players

reload(sys)
sys.setdefaultencoding('utf-8')


def analise(directory):
    files = collect_files(directory)
    leagues = []
    for file in files:
        players = load_league(file)
        leagues.append(league)

    build_workbook(leagues)


def players_from_leagues(leagues):
    players = set()
    for league in leagues:
        for player in league['players'].keys():
            players.add(player)
    players = sorted(players)
    return players


def build_workbook(leagues):
    wb = load_workbook(file)

def collect_files(directory):
    files_to_match = os.path.join(directory, "*.xlsx")
    files = glob.glob(files_to_match)
    files.sort(key=natural_key)
    return files


def load_league(file):
    _, filename = os.path.split(file)
    (shortname, extension) = os.path.splitext(filename)
    league_num = int(shortname.split(" ")[-1])
    print("Processing {}".format(shortname))
    league = {}
    wb = load_workbook(file)
    for sheet in wb:
        week = int(sheet.title)
        for row in sheet.iter_rows(row_offset=2):
            player = row[1].value
            if player is None:
                continue
            player = str(player)
            player = player.lstrip().rstrip()
            if player not in league:
                league[player] = {}
            if week in league[player]:
                league[player][week] += 1
            else:
                league[player][week] = 1

    return {'league': league_num, 'players': league}

def natural_key(string_):
    """See http://www.codinghorror.com/blog/archives/001018.html"""
    return [int(s) if s.isdigit() else s for s in re.split(r'(\d+)', string_)]


for filename in files:
    _, league = load_players(filename, players_wb)
    leagues.append(league)

players = set()
for league in leagues:
    for player in league['players'].keys():
        players.add(player)
players = sorted(players)

ws = players_wb.worksheets[0]
ws.title = "Players"
ws['A2'].value = "Name"
ws['B2'].value = "Biggest Gap"
ws['C2'].value = "Weeks Played"
ws['A2'].fill = PatternFill(patternType='solid', fgColor=Color("84D6F0"))
ws['B2'].fill = PatternFill(patternType='solid', fgColor=Color("84D6F0"))
ws['C2'].fill = PatternFill(patternType='solid', fgColor=Color("84D6F0"))

sorted_leagues = sorted(leagues, key=itemgetter('league'))

center_alignment = Alignment(horizontal='center')
last_column = None
for col, league in enumerate(sorted_leagues, start=1):
    league_num = league['league']
    league_header_cell = ws.cell(row=2, column=col+3)
    league_header_cell.value = league_num
    league_header_cell.fill = PatternFill(patternType='solid', fgColor=Color("84D6F0"))
    league_header_cell.alignment = center_alignment

    ws.column_dimensions[league_header_cell.column].width = 3
    last_column = column_index_from_string(league_header_cell.column)

ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=last_column)
league_title_cell = ws.cell(row=1, column=4)
league_title_cell.value = "Leagues"
league_title_cell.alignment = center_alignment
league_title_cell.fill = PatternFill(patternType='solid', fgColor=Color("FCB082"))

yellow_fill = PatternFill(patternType='solid', fgColor=Color("E6F2A0"))
red_fill = PatternFill(patternType='solid', fgColor=Color("D49093"))
blue_fill = PatternFill(patternType='solid', fgColor=Color("8290FC"))
pink_fill = PatternFill(patternType='solid', fgColor=Color("FFDCFF"))
medium_style = Side(style='medium', color=Color("A600BC"))
medium_border = Border(left=medium_style, right=medium_style, top=medium_style, bottom=medium_style)

for num, player in enumerate(players, start=2):
    started_bowling = False
    row = 1 + num
    player_cell = ws.cell(row=row, column=1)
    gap_cell = ws.cell(row=row, column=2)
    player_cell.value = player
    player_clash = False
    week_count = 0
    league_gaps = []
    league_gap = 0
    league_gap_threshold = 5
    for col, league in enumerate(sorted_leagues, start=1):
        league_cell = ws.cell(row=row, column=col+3)
        league_num = league['league']
        if player not in league['players']:
            if started_bowling:
                league_gap += 1
            continue
        if started_bowling and (league_gap >= league_gap_threshold):
            for x in range(league_gap, 0, -1):
                cell = ws.cell(row=row, column=col+3-x)
                cell.fill = pink_fill
            gap_cell.fill = pink_fill
        league_gaps.append(league_gap)
        league_gap = 0
        started_bowling = True
        weeks = league['players'][player]
        played_weeks = sum(weeks.values())
        league_cell.value = played_weeks
        league_cell.alignment = center_alignment
        max_key = max(weeks, key=weeks.get)
        if weeks[max_key] == 1:
            league_cell.fill = yellow_fill
        elif weeks[max_key] > 1:
            league_cell.fill = red_fill
            player_clash = True

        week_count += played_weeks
    gap_cell.value = max(league_gaps)
    week_count_cell = ws.cell(row=row, column=3)
    week_count_cell.value = week_count
    if player_clash:
        week_count_cell.fill = red_fill

players_wb.save('leagueplayers.xlsx')
