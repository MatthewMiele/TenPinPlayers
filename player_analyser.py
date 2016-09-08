import os
import glob
from openpyxl import load_workbook
from collections import OrderedDict
from natsort import natsorted

from workbook import PlayerWorkbook


def analyse(directory, outputfilename, leaguegapthreshold):
    files = collect_files(directory)
    leagues = []
    for file in files:
        league_players = load_league_from_file(file)
        leagues.append(league_players)
    pw = PlayerWorkbook(league_gap_threshold=leaguegapthreshold)
    workbook = pw.build_workbook(leagues)
    workbook.save('{}.xlsx'.format(outputfilename))


def collect_files(directory):
    files_to_match = os.path.join(directory, "*.xlsx")
    files = glob.glob(files_to_match)
    files = natsorted(files)
    return files


def load_league_from_file(file):
    _, filename = os.path.split(file)
    (shortname, extension) = os.path.splitext(filename)
    league_num = int(shortname.split(" ")[-1])
    print("Processing {}".format(league_num))
    players = {}
    wb = load_workbook(file)
    for sheet in wb:
        week = int(sheet.title)
        for row in sheet.iter_rows(row_offset=2):
            player = row[1].value
            if player is None:
                continue
            player = str(player)
            player = player.upper().lstrip().rstrip()
            if player not in players:
                players[player] = {}
            if week in players[player]:
                players[player][week] += 1
            else:
                players[player][week] = 1

    players = OrderedDict(sorted(players.items()))
    no_of_weeks = len(wb.worksheets)
    return {
        'league': league_num,
        'no_of_weeks': no_of_weeks,
        'players': players
    }
