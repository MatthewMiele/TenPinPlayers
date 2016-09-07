from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.utils import column_index_from_string
from collections import Counter, OrderedDict
import os
import operator
from timeit import default_timer as timer


def load_players(file, workbook):
    wb = load_workbook(file)
    _, filename = os.path.split(file)
    (shortname, extension) = os.path.splitext(filename)
    league_num = int(shortname.split(" ")[-1])

    print("Processing {}".format(shortname))
    league = {}
    for sheet in wb:
        week = int(sheet.title)
        for row in sheet.iter_rows(row_offset=2):
            player = row[1].value
            if player is None:
                continue
            player = str(player)
            player = player.lstrip().rstrip()
            if player not in league:
                league[player] = {}
            if week in league[player]:
                league[player][week] += 1
            else:
                league[player][week] = 1

    ws = workbook.create_sheet()
    ws['A2'].value = "Name"
    ws['B2'].value = "Weeks Played"

    ws['A2'].fill = PatternFill(patternType='solid', fgColor=Color("84D6F0"))
    ws['B2'].fill = PatternFill(patternType='solid', fgColor=Color("84D6F0"))

    center_alignment = Alignment(horizontal='center')
    last_column = None
    for num in range(1, len(wb.worksheets) + 1):
        week_header_cell = ws.cell(row=2, column=num+2)
        week_header_cell.value = num
        week_header_cell.fill = PatternFill(patternType='solid', fgColor=Color("84D6F0"))
        week_header_cell.alignment = center_alignment
        ws.column_dimensions[week_header_cell.column].width = 3
        last_column = column_index_from_string(week_header_cell.column)

    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=last_column)
    week_title_cell = ws.cell(row=1, column=3)
    week_title_cell.value = "Weeks"
    week_title_cell.fill = PatternFill(patternType='solid', fgColor=Color("FCB082"))
    week_title_cell.alignment = center_alignment


    league = OrderedDict(sorted(league.items()))

    yellow_fill = PatternFill(patternType='solid', fgColor=Color("E6F2A0"))
    red_fill = PatternFill(patternType='solid', fgColor=Color("D49093"))

    for num, (player, weeks) in enumerate(league.iteritems(), start=2):
        row = num + 1
        ws.cell(row=row, column=1).value = player
        ws.cell(row=row, column=2).value = sum(weeks.values())
        for week, count in weeks.iteritems():
            weekcell = ws.cell(row=row, column=week+2)
            weekcell.value = count
            weekcell.alignment = center_alignment
            if count == 1:
                weekcell.fill = yellow_fill
            elif count > 1:
                weekcell.fill = red_fill

    ws.title = shortname
    return ws, {'league': league_num, 'players': league}
